import os
import re
import glob
from datetime import datetime
import warnings
import pandas as pd
import sys

# ===== 可調參數 =====
DATE_COL_CANDIDATES = ["憑單日期", "憑單日", "憑單日期(西元)", "單據日期"]  # 統一輸出成第一個名稱
# 可能的金額欄位（會轉成數字，Excel 才會自動加總）
AMOUNT_COL_CANDIDATES = [
    "本幣貨款金額",
    "本幣金額",
    "金額(本幣)",
    "金額本幣",
    "金額",
]

READ_ALL_SHEETS = True
INCLUDE_SUBDIRS = False
EXCEL_PATTERNS = ["*.xlsx", "*.xls"]
SKIP_PREFIXES = ("~$",)
SORT_ASC = True
DATE_FMT_DEFAULT = "yyyy/mm/dd"     # Excel 顯示格式（年月日）
DATE_COLUMN_WIDTH = 12
DEFAULT_COL_WIDTH = 8
MAX_EXCEL_STRLEN = 32767

# 想要偵測的「正式欄位列」關鍵欄位
HEADER_KEYS = [
    "憑單日期",
    "憑單單號",
    "單據日期",
    "廠商代號",
    "廠商簡稱",
    "付款條件代號",
    "付款條件名稱",
    "預計付款日",
    "預計兌現日",
]

# Excel 禁用/不合法 XML 字元（保留 \t \n \r）
ILLEGAL_RE = re.compile(
    r"["                     # 開頭 [
    r"\x00-\x08\x0B\x0C\x0E-\x1F"   # C0 控制字元（排除 \t \n \r）
    r"\x7F-\x84\x86-\x9F"           # C1 控制字元
    r"\uD800-\uDFFF"                # 代理對區段（孤立代理）
    r"\uFDD0-\uFDEF"                # 非字元碼點區段
    r"\uFFFE\uFFFF"                 # 非字元
    r"]"                     # 結尾 ]
)

# ===== 基準與輸出 =====
if getattr(sys, "frozen", False):
    # 被 PyInstaller 打包成 EXE 時，抓 EXE 本身所在資料夾
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # 平常用 python merge_all_data.py 跑時，抓 .py 所在資料夾
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.join(BASE_DIR, "total")
os.makedirs(OUTPUT_DIR, exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"TOTAL_{timestamp}.xlsx")

# 自動排除的目錄 / 檔名（避免把輸出或報表再吃回去）
EXCLUDE_DIR_NAMES = {
    os.path.basename(OUTPUT_DIR),  # 'total'
    "out",
    "compare",
    "__pycache__",
    "venv",
    ".venv",
    ".git",
}
EXCLUDE_DIR_CONTAINS = ["backup", "_bak", "test"]

EXCLUDE_FILE_PATTERNS = [
    "TOTAL_*.xlsx",         # 舊的總表
    "summary_*.xlsx",       # 圓餅圖彙總
    "TERMS_COMPARE_*.xlsx", # 付款條件比對報表
]

# ===== 基本工具 =====
def list_immediate_subdirs(path):
    # 排除輸出目錄 & 特定資料夾
    results = []
    for d in os.listdir(path):
        full = os.path.join(path, d)
        if not os.path.isdir(full):
            continue
        name = d
        if name in EXCLUDE_DIR_NAMES:
            continue
        low = name.lower()
        if any(key in low for key in EXCLUDE_DIR_CONTAINS):
            continue
        results.append(name)
    return sorted(results)

def list_excels(folder):
    files = []
    for pat in EXCEL_PATTERNS:
        if INCLUDE_SUBDIRS:
            files += glob.glob(os.path.join(folder, "**", pat), recursive=True)
        else:
            files += glob.glob(os.path.join(folder, pat))

    # 排除暫存/系統前綴 & 排除特定樣式的檔名
    filtered = []
    for f in files:
        base = os.path.basename(f)
        if base.startswith(SKIP_PREFIXES):
            continue
        if any(glob.fnmatch.fnmatch(base, patt) for patt in EXCLUDE_FILE_PATTERNS):
            continue
        filtered.append(f)
    return sorted(filtered)

def coerce_date_series(s: pd.Series) -> pd.Series:
    """
    穩定日期解析：
    1) 先把空白 / NaN 正規化
    2) 日期往下填滿（ffill），處理「只有第一列有日期，下面明細空白」的情況
    3) 依序嘗試不同日期格式與 Excel 序號
    """
    tmp = s.astype(str).str.strip()
    tmp = tmp.replace({"": pd.NA, "nan": pd.NA, "NaT": pd.NA})

    # 日期往下填滿
    tmp = tmp.ffill()

    dt = pd.to_datetime(tmp, errors="coerce", format="%Y/%m/%d")
    mask = dt.isna()
    if mask.any():
        dt2 = pd.to_datetime(tmp[mask], errors="coerce", format="%Y-%m-%d")
        dt.loc[mask] = dt2

    mask = dt.isna() & pd.to_numeric(tmp, errors="coerce").notna()
    if mask.any():
        nums = pd.to_numeric(tmp[mask], errors="coerce")
        dt2 = pd.to_datetime(nums, unit="D", origin="1899-12-30", errors="coerce")
        dt.loc[mask] = dt2

    return dt.dt.floor("D")

def find_latest_template():
    candidates = glob.glob(os.path.join(OUTPUT_DIR, "TOTAL_*.xlsx"))
    candidates = [p for p in candidates if os.path.abspath(p) != os.path.abspath(OUTPUT_FILE)]
    if not candidates:
        return None
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]

def _clean_cell_value(v):
    if pd.isna(v):
        return v
    s = str(v)
    s = ILLEGAL_RE.sub("", s)
    if len(s) > MAX_EXCEL_STRLEN:
        s = s[:MAX_EXCEL_STRLEN]
    return s

def read_template_header(template_path):
    """僅讀上一份 TOTAL 的表頭欄名（不複製樣式）。"""
    if not template_path:
        return None
    try:
        from openpyxl import load_workbook
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
        wb = load_workbook(template_path, read_only=True, data_only=True)
        if "TOTAL" not in wb.sheetnames:
            return None
        ws = wb["TOTAL"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header = [c for c in header if c is not None]
        return header or None
    except Exception:
        return None

# ===== 依欄位名稱自動找出真正表頭列 =====
def extract_table_from_sheet(df_raw: pd.DataFrame):
    """
    df_raw 是 header=None 讀進來的工作表
    會從上往下找含有 HEADER_KEYS 的那一列作為欄位列
    找到後，該列以下才視為資料
    """
    header_row = None
    nrows = len(df_raw)

    for i in range(nrows):
        row = df_raw.iloc[i]
        values = [str(v).strip() for v in row if not pd.isna(v) and str(v).strip() != ""]
        if not values:
            continue
        # 計算這列裡有幾個欄位名稱命中 HEADER_KEYS
        hit = sum(1 for key in HEADER_KEYS if key in values)
        if hit >= 3:  # 至少對到 3 個就視為正式欄位列
            header_row = i
            break

    if header_row is None:
        return None

    header = df_raw.iloc[header_row]
    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = header

    # 移除整列空白
    df = df.dropna(how="all")

    # 移除完全是 Unnamed 且內容全空的欄位
    mask_keep = []
    for col in df.columns:
        col_str = str(col)
        if col_str.startswith("Unnamed"):
            if df[col].notna().any():
                mask_keep.append(True)
            else:
                mask_keep.append(False)
        else:
            mask_keep.append(True)
    df = df.loc[:, mask_keep]

    return df

# ===== 讀檔與彙整 =====
records, file_log, errors = [], [], []

print(f"Base dir: {BASE_DIR}")
subdirs = list_immediate_subdirs(BASE_DIR)
print(f"偵測子資料夾（會合併）：{subdirs}")

for sub in subdirs:
    subdir_path = os.path.join(BASE_DIR, sub)
    excel_files = list_excels(subdir_path)
    if not excel_files:
        continue

    for fpath in excel_files:
        # 雙重保險：避免任何位於 OUTPUT_DIR 的檔案被吃回
        if os.path.commonpath([os.path.abspath(fpath), os.path.abspath(OUTPUT_DIR)]) == os.path.abspath(OUTPUT_DIR):
            continue

        try:
            # 用 header=None 讀，讓 extract_table_from_sheet 自己找表頭列
            if READ_ALL_SHEETS:
                xls = pd.read_excel(fpath, sheet_name=None, header=None)
                items = xls.items()
            else:
                df_raw = pd.read_excel(fpath, header=None)
                items = [("Sheet1", df_raw)]

            for sht, df_raw in items:
                if df_raw is None or df_raw.empty:
                    continue

                df = extract_table_from_sheet(df_raw)
                if df is None or df.empty:
                    errors.append({"file": f"{fpath}::{sht}", "error": "找不到欄位列"})
                    continue

                # 找日期欄
                date_col = next((c for c in DATE_COL_CANDIDATES if c in df.columns), None)
                if date_col is None:
                    if len(df.columns) >= 2:
                        date_col = df.columns[1]
                    else:
                        continue

                dt = coerce_date_series(df[date_col])
                keep = dt.notna()
                if not keep.any():
                    continue

                df2 = df.loc[keep].copy()

                # 清洗欄名
                df2.columns = pd.Index([_clean_cell_value(c) if c is not None else c for c in df2.columns])

                # 統一日期欄名
                std_date_name = DATE_COL_CANDIDATES[0]
                df2[std_date_name] = dt[keep]

                # 移除 Unnamed 欄
                df2 = df2.loc[:, ~df2.columns.astype(str).str.startswith("Unnamed")]

                # 清洗內容（先不動金額欄位，後面統一處理）
                for col in df2.columns:
                    if col in AMOUNT_COL_CANDIDATES:
                        continue
                    if df2[col].dtype == "object":
                        df2[col] = df2[col].map(_clean_cell_value)

                # 附加來源資訊
                df2["source_folder"] = sub
                df2["source_file"] = os.path.basename(fpath)
                df2["sheet_name"] = sht

                records.append(df2)
                file_log.append({"folder": sub, "file": fpath, "sheet": sht, "rows": len(df2)})

        except Exception as e:
            errors.append({"file": fpath, "error": str(e)})

# ===== 欄序對齊、排序、日期只保留年月日 =====
template_path = find_latest_template()
template_cols = read_template_header(template_path)

if records:
    total_df = pd.concat(records, ignore_index=True)

    # 二次保險清洗（先清欄名）
    total_df.columns = pd.Index([_clean_cell_value(c) if c is not None else c for c in total_df.columns])

    # 1) 非金額欄位：清掉非法字元
    for col in total_df.columns:
        if col in AMOUNT_COL_CANDIDATES:
            continue
        if total_df[col].dtype == "object":
            total_df[col] = total_df[col].map(_clean_cell_value)

    # 2) 金額欄位：轉成數字（Excel 才會自動加總）
    for col in AMOUNT_COL_CANDIDATES:
        if col in total_df.columns:
            s = total_df[col].astype(str).str.replace(",", "", regex=False).str.strip()
            s = s.replace({"": pd.NA, "nan": pd.NA, "NaN": pd.NA})
            total_df[col] = pd.to_numeric(s, errors="coerce")

    # 欄序對齊（僅依上一份 TOTAL 的表頭順序，無樣式）
    if template_cols:
        if DATE_COL_CANDIDATES[0] not in total_df.columns and any(x == DATE_COL_CANDIDATES[0] for x in template_cols):
            total_df[DATE_COL_CANDIDATES[0]] = pd.NaT
        aligned_cols = template_cols + [c for c in total_df.columns if c not in template_cols]
        for c in template_cols:
            if c not in total_df.columns:
                total_df[c] = pd.NA
        total_df = total_df[aligned_cols]

    # 排序（先依日期，再依來源：資料夾 / 檔案 / 工作表）
    sort_cols = [DATE_COL_CANDIDATES[0]] + [
        c for c in ["source_folder", "source_file", "sheet_name"] if c in total_df.columns
    ]
    total_df = total_df.sort_values(by=sort_cols, ascending=SORT_ASC).reset_index(drop=True)

    # 只保留「年月日」
    if DATE_COL_CANDIDATES[0] in total_df.columns:
        total_df[DATE_COL_CANDIDATES[0]] = pd.to_datetime(
            total_df[DATE_COL_CANDIDATES[0]],
            errors="coerce"
        ).dt.date
else:
    total_df = pd.DataFrame(columns=[DATE_COL_CANDIDATES[0]])

# ===== 穩定輸出（XlsxWriter，一次成型）=====
with pd.ExcelWriter(OUTPUT_FILE, engine="xlsxwriter") as writer:
    total_df.to_excel(writer, index=False, sheet_name="TOTAL")
    (pd.DataFrame(file_log) if file_log else pd.DataFrame(columns=["folder", "file", "sheet", "rows"])) \
        .to_excel(writer, index=False, sheet_name="SUMMARY")
    if errors:
        pd.DataFrame(errors).to_excel(writer, index=False, sheet_name="ERRORS")

    wb  = writer.book
    ws  = writer.sheets["TOTAL"]

    headers = list(total_df.columns)

    # 基本欄寬（避免 ####）
    ws.set_column(0, len(headers) - 1, DEFAULT_COL_WIDTH)

    # 日期欄格式 + 欄寬
    if DATE_COL_CANDIDATES[0] in headers:
        date_col_idx0 = headers.index(DATE_COL_CANDIDATES[0])  # 0-based
        date_fmt = wb.add_format({"num_format": DATE_FMT_DEFAULT})
        ws.set_column(date_col_idx0, date_col_idx0, DATE_COLUMN_WIDTH, date_fmt)

    # 凍結首列
    ws.freeze_panes(1, 0)

    # AutoFilter 覆蓋資料範圍（含表頭）
    nrows, ncols = (len(total_df.index) + 1, len(headers))
    ws.autofilter(0, 0, nrows - 1, ncols - 1)

print(f"已輸出：{OUTPUT_FILE}\n（欄序來源：{template_path if template_cols else '無，直接以新資料欄序'}）")
